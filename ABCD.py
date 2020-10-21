
# this is a python script, which wrote to get restaurants' details from "Trip Advisor" web site
# still programming


import csv
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

PATH = "C:\python file\chromedriver.exe"

driver = webdriver.Chrome(PATH)

driver.get("https://www.google.com/")

obj = openpyxl.load_workbook("tokyorestu.xlsx")

# print(obj)

sheet_obj = obj.active
# print(sheet_obj)

m_row = sheet_obj.max_row

for i in range(1, m_row):
    cell_obj = sheet_obj.cell(row=i, column=1)
    restaurantname = cell_obj.value
    driver.get(
        "https://www.tripadvisor.jp/Restaurant_Review-g1066443-d12987949-Reviews-Burger_Milkshake_Crane-Chiyoda_Tokyo_Tokyo_Prefecture_Kanto.html")
    search = driver.find_element_by_class_name("_3qLQ-U8m")
    search.send_keys(restaurantname)
    search.send_keys(Keys.RETURN)

    try:
        main = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "location-meta-block"))
        )

        element = main.find_element_by_class_name("review_count")

        # print(element.text)
        linkmain = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.LINK_TEXT, element.text))
        )

        linkmain.click()

        tabs = driver.window_handles
        for tab in tabs:
            driver.switch_to.window(tab)

        asi = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "page "))
        )
        address = asi.find_element_by_class_name("_2saB_OSe")

        telephonenumber = asi.find_element_by_x_path("//a/span[1]/span[2]")

        print(address.text, telephonenumber.text)


    except:
        print(" exception occurred")
