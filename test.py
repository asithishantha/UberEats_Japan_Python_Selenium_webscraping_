import csv
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import csv

# here, i give path to the web driver(for this project, i used chrome web browser).
# you can download any web driver relevant to your web browser and the version and give path to the location
# for chrome driver : https://chromedriver.chromium.org/

PATH = "C:\python file\chromedriver.exe"

url = " "
driver = webdriver.Chrome(PATH)

driver.get("https://www.google.com/")
# website you wish to scrape

obj = openpyxl.load_workbook("nagoya.xlsx")
# load the excel sheet you wish to read. In this program, we are getting the restaurant name from the excel sheet
# search each restaurants' name in google to find the address and the mobile number.

# Note: to run this program, you must have the input excel sheet that include information to read.
# In my case "restaurants' name"


# print(obj)

sheet_obj = obj.active
# print(sheet_obj)

m_row = sheet_obj.max_row

for i in range(1, m_row):
    cell_obj = sheet_obj.cell(row=i, column=1)
    restaurantname = cell_obj.value  # take the restaurant name in first cell
    driver.get("https://www.google.com/")
    search = driver.find_element_by_name("q")  # google search box element name by name, we can use class name,
    # tag name and any other locate element methods.
    search.send_keys(restaurantname)  # search the restaurant name
    search.send_keys(Keys.RETURN)  # press search

    try:

        # we can use any locating method to locate elements. Here i used get_element_by_x_path to locate elements.
        # there are many locating methods. get element by id, get element by class name, get element by tag name.
        # please use this video tutorial for more details.
        # https://www.youtube.com/watch?v=mQ7-mPJYJ5A&t=973s

        number2 = driver.find_element_by_xpath("//span[@class='LrzXr zdqRlf kno-fv']/a[1]/span[1]")
        mobnumber = number2.text
        # print(mobnumber)


    except:
        # if the program couldn't find the targeted element, print an exception message
        mobnumber = "mobile number exception occurred"

    try:
        mobmain = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "ifM9O"))
        )  # 15 second waiting until load the page correctly

        address = mobmain.find_element_by_class_name("LrzXr")
        # here, search the address element by its class name, use F12 to enter the inspect and find the location of
        # element(ex: address, mobile number, name , email you  wish to scrape ) and identify a class name or tag name
        # related to that element. you can use x_path also . if you don't know how to write x path,
        # you can use google chrome extension that will generate the xpath .
        # for more details : https://www.youtube.com/watch?v=VvZEsZ3cGmc

        # print(address.text)
        address = address.text


    except:
        address = "address exception occurred"

    resturant = {

        'name': restaurantname,
        'address': address,
        'number': mobnumber
    }


    def write_csv(resturant):
        with open('nagoya_res.csv', 'a', encoding='utf8') as csvfile:
            writer = csv.writer(csvfile)

            row = [resturant['name'], resturant['address'], resturant['number'], url]

            writer.writerow(row)


    write_csv(resturant)
    # write to the csv file. please give the csv file name accordingly, you wish to write the scraped data
    print(resturant)
